# genau auf das sheet von 2023 mit den Arten von Stand 27.07.2023 angepasst
import openpyxl
from openpyxl.styles import numbers, PatternFill

print("Voraussetzung: Arbeitsverzeichnis ist in der IDE geöffnet")
xlsxDIR = str(input("Pfad der XLSX eingeben [Format: Dateiname.xlsx] "))
xlsxSHEET = str(input("Name des Tabellensheets eingeben "))
xlsxSAVE = str(input("Speicherpfad der XLSX eingeben [Format: Dateiname.xlsx] "))

FirstCell = str(input("Name der ersten Zelle mit Daten eingeben (z.B. F4) "))
LastCell = str(input("Name der letzten Zelle mit Daten eingeben (z.B. IP244) "))

Moss = input("\n Bitte alle Spaltennamen der Moose angeben (Format: AR, AS, ...) ")
MossList = Moss.split(",")
print(MossList)

fileXLSX = openpyxl.load_workbook(xlsxDIR)
sheet = fileXLSX[xlsxSHEET]
for col in sheet[FirstCell : LastCell]:
    for row in col:
#        print(f"{row.value} - {row.coordinate} - {type(row.value)}")
        coord = openpyxl.utils.cell.coordinate_to_tuple(row.coordinate)
        col_lett = openpyxl.utils.cell.get_column_letter(coord[1])
#        print(f"{coord[1]} - {col_lett}")
        if row.value == None:
                #sheet.cell(coord[0],coord[1]).value = 0
            sheet.cell(coord[0], coord[1]).value = ""
        elif type(row.value) == str:
            if row.value == "x":
                sheet.cell(coord[0], coord[1]).number_format = numbers.FORMAT_TEXT
                sheet.cell(coord[0], coord[1]).value = "b"
            elif row.value == "x+":
                sheet.cell(coord[0], coord[1]).number_format = numbers.FORMAT_TEXT
                sheet.cell(coord[0], coord[1]).value = 3
            elif row.value == "x++":
                sheet.cell(coord[0], coord[1]).number_format = numbers.FORMAT_TEXT
                sheet.cell(coord[0], coord[1]).value = 4
            #   sheet.cell(coord[0],coord[1]).value = 0
            #   sheet.cell(coord[0],coord[1]).value = int(row.value)
            #   print(f"{row.value} - {col.coordinate}")
        elif row.value == 1:
            sheet.cell(coord[0], coord[1]).number_format = numbers.FORMAT_TEXT
            sheet.cell(coord[0], coord[1]).value = "R"
            # + und 1 sind manuell zu behandeln, da keine Unterscheidung mit python mgl.
        elif int(row.value) > 1 and int(row.value) < 5:
            if col_lett in Moss:
               sheet.cell(coord[0], coord[1]).value = "2m"
                #sheet.cell(coord[0], coord[1]).fill = PatternFill(start_color='AAAA9999', end_color='AAAA9999', fill_type='solid')
            else: 
                sheet.cell(coord[0], coord[1]).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        elif int(row.value) >= 5 and int(row.value) <= 12:  # 12 ist eig 12,5
                sheet.cell(coord[0], coord[1]).number_format = numbers.FORMAT_TEXT
                sheet.cell(coord[0], coord[1]).value = "2a"
        elif int(row.value) > 12 and int(row.value) < 25:  # 12 ist eig 12,5
                sheet.cell(coord[0], coord[1]).number_format = numbers.FORMAT_TEXT
                sheet.cell(coord[0], coord[1]).value = "2b"
        elif int(row.value) >= 25 and int(row.value) < 50:
            sheet.cell(coord[0], coord[1]).value = 3
        elif int(row.value) >= 50 and int(row.value) < 75:
                sheet.cell(coord[0], coord[1]).value = 4
        elif int(row.value) >= 75 and int(row.value) < 100:
                sheet.cell(coord[0], coord[1]).value = 5      
#        print(col.coordinate)
fileXLSX.save(xlsxSAVE)
print("Saved")